# -*- coding: utf-8 -*-
"""把一家企业的多个企查查 JSON 快照整合为一份便于阅读的 Excel。

输出文件名:qcc_{企业名称}_{YYYYmmdd_HHMMSS}.xlsx,两列(指标名 / 指标值)。
指标全集与取值口径取自 00-对齐/05-企查查接口数据法务确认表.xlsx 的
「字段对照与实测数据」sheet(19 个主体回填 + 23 个风险项,共 42 项),
口径与 01-poc/build_legal_review_xlsx.py 保持一致,仅改成两列布局。

两种调用来源:
  - aggregator.fetch_profile 跑完直接用内存 payloads 调 build_report(自动生成)
  - query.py export <企业名> 用 load_payloads 从 snapshots/ 读最新快照后调 build_report(脱机,零成本)
"""
from __future__ import annotations
import glob
import json
import os
import time
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", start_color="D9E1F2")
SECTION_FONT = Font(name=FONT, bold=True, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---- 取值辅助(移植自 01-poc/build_legal_review_xlsx.py,口径不变) ----

def _risk_cell(value: Any) -> str:
    if value is None:
        return "无记录"
    if isinstance(value, dict):
        tc = value.get("TotalCount")
        if tc is not None:
            try:
                n = int(tc)
                return "无记录" if n == 0 else "%d条" % n
            except (TypeError, ValueError):
                pass
        return "有记录"
    if isinstance(value, list):
        return "%d条" % len(value) if value else "无记录"
    return str(value)


def _list_summary(items: List[dict], fmt: Callable[[dict], str]) -> str:
    if not items:
        return "无记录"
    out = "; ".join(fmt(x) for x in items[:3])
    return out + (" 等" if len(items) > 3 else "")


def _supp(ctx: dict, code: int) -> dict:
    return (ctx.get("_supp") or {}).get(str(code)) or {}


def _supp_status(ctx: dict, code: int) -> str:
    return str(_supp(ctx, code).get("Status", ""))


def supp_856(ctx: dict) -> str:
    resp = _supp(ctx, "856")
    if str(resp.get("Status", "")) != "200":
        return "未开通(214)"
    vr = (resp.get("Result") or {}).get("VerifyResult")
    return {1: "一致(VerifyResult=1)", 0: "公司编号有误(0)",
            2: "企业名称不一致(2)", 3: "法定代表人不一致(3)"}.get(vr, "未知(%s)" % vr)


def _supp_list(ctx: dict, code: int, fmt: Callable[[dict], str]) -> str:
    resp = _supp(ctx, code)
    if str(resp.get("Status", "")) != "200":
        return "未开通(214)"
    result = resp.get("Result") or {}
    if result.get("VerifyResult") != 1:
        return "无记录"
    data = result.get("Data") or []
    total = (resp.get("Paging") or {}).get("TotalRecords")
    head = "共%s条; " % total if total is not None else ""
    return head + _list_summary(data, fmt)


def supp_739(ctx: dict) -> str:
    return _supp_list(ctx, "739", lambda x: "%s(%s)" % (x.get("AddDate"), x.get("AddReason")))


def supp_887(ctx: dict) -> str:
    return _supp_list(ctx, "887", lambda x: "%s %s(%s)" % (x.get("JudgeDate"), x.get("CaseReason"), x.get("CaseNo")))


def supp_888(ctx: dict) -> str:
    return _supp_list(ctx, "888", lambda x: "%s %s(%s)" % (x.get("CourtTime"), x.get("CaseReason"), x.get("CaseNo")))


def supp_889(ctx: dict) -> str:
    return _supp_list(ctx, "889", lambda x: "%s %s(%s)" % (x.get("PunishDate"), x.get("CaseReason"), x.get("CaseNo")))


def _f_subject(key: str) -> Callable[[dict], str]:
    return lambda d: str(d.get(key) or "无记录")


def _f_nested(path: str) -> Callable[[dict], str]:
    keys = path.split(".")

    def ext(d: dict) -> str:
        v: Any = d
        for k in keys:
            v = (v or {}).get(k) if isinstance(v, dict) else None
        return str(v) if v not in (None, "") else "无记录"
    return ext


def _f_actual_controller(d: dict) -> str:
    return _list_summary(d.get("ActualControllerList"),
                         lambda x: "%s(%s)" % (x.get("Name"), x.get("FinalBenefitPercent") or "?"))


def _f_beneficiary(d: dict) -> str:
    return _list_summary(d.get("BeneficiaryList"),
                         lambda x: "%s(%s)" % (x.get("Name"), x.get("FinalBenefitPercent") or "?"))


def _f_partners(d: dict) -> str:
    return _list_summary(d.get("PartnerList"),
                         lambda x: "%s %s" % (x.get("StockName"), x.get("StockPercent")))


def _f_employees(d: dict) -> str:
    return _list_summary(d.get("EmployeeList"),
                         lambda x: "%s(%s)" % (x.get("Name"), x.get("Job")))


def _f_tax_credit(d: dict) -> str:
    return _list_summary(d.get("TaxCreditList"),
                         lambda x: "%s年 %s级" % (x.get("Year"), x.get("Level")))


def _f_change_recent(d: dict) -> str:
    cl = d.get("ChangeList") or []
    recent = [c for c in cl if str(c.get("ChangeDate", "")) >= "2025-01-01"]
    if not recent:
        return "近1年无变更记录"
    return _list_summary(recent, lambda x: "%s %s" % (x.get("ChangeDate"), x.get("ProjectName")))


def _f_area(d: dict) -> str:
    a = d.get("Area") or {}
    return "%s%s%s" % (a.get("Province") or "", a.get("City") or "", a.get("County") or "") or "无记录"


def _f_term(d: dict) -> str:
    return "%s 至 %s" % (d.get("TermStart") or "?", d.get("TermEnd") or "无期限")


def _f_revoke(d: dict) -> str:
    return "无记录" if d.get("RevokeInfo") is None else "有记录"


# ---- 指标全集:(分组, 指标名, 取值函数) ----
# 分组沿用 05 表「风险类别」列;指标名沿用 05 表「需求信息项」列原文。

INDICATORS: List[tuple] = [
    # 主体回填(19)
    ("主体回填", "企业全称", _f_subject("Name")),
    ("主体回填", "统一社会信用代码", _f_subject("CreditCode")),
    ("主体回填", "法定代表人", _f_subject("OperName")),
    ("主体回填", "注册资本", _f_subject("RegistCapi")),
    ("主体回填", "成立日期", _f_subject("StartDate")),
    ("主体回填", "营业期限", _f_term),
    ("主体回填", "企业类型", _f_subject("EconKind")),
    ("主体回填", "登记状态", _f_subject("Status")),
    ("主体回填", "注册地址", _f_subject("Address")),
    ("主体回填", "联系电话", _f_nested("ContactInfo.Tel")),
    ("主体回填", "联系邮箱", _f_nested("ContactInfo.Email")),
    ("主体回填", "经营范围", _f_subject("Scope")),
    ("主体回填", "所属地区", _f_area),
    ("主体回填", "实控人", _f_actual_controller),
    ("主体回填", "受益所有人", _f_beneficiary),
    ("主体回填", "股东", _f_partners),
    ("主体回填", "主要人员", _f_employees),
    ("主体回填", "纳税信用等级", _f_tax_credit),
    ("一致性校验", "三要素(代码+名称+法人)一致性", supp_856),
    # 经营状态(2)
    ("经营状态", "存续/注销/吊销/迁出", _f_subject("Status")),
    ("经营状态", "吊销注销信息", _f_revoke),
    # 工商异常(3)
    ("工商异常", "经营异常名录", lambda d: _risk_cell(d.get("Exception"))),
    ("工商异常", "经营异常明细(739专项)", supp_739),
    ("工商异常", "严重违法失信名单", lambda d: _risk_cell(d.get("SeriousIllegal"))),
    # 司法风险(9)
    ("司法风险", "失信被执行人", lambda d: _risk_cell(d.get("ShiXin"))),
    ("司法风险", "被执行人", lambda d: _risk_cell(d.get("ZhiXing"))),
    ("司法风险", "限制高消费", lambda d: _risk_cell(d.get("Sumptuary"))),
    ("司法风险", "破产重整", lambda d: _risk_cell(d.get("Bankruptcy"))),
    ("司法风险", "股权冻结", lambda d: _risk_cell(d.get("EquityFreeze"))),
    ("司法风险", "司法拍卖", lambda d: _risk_cell(d.get("JudicialSale"))),
    ("司法风险", "涉诉-裁判文书明细", supp_887),
    ("司法风险", "涉诉-开庭公告明细", supp_888),
    ("司法风险", "涉诉-立案信息明细", supp_889),
    # 行政处罚(2)
    ("行政处罚", "行政处罚记录", lambda d: _risk_cell(d.get("AdminPenalty"))),
    ("行政处罚", "环保处罚", lambda d: _risk_cell(d.get("EnvPunishment"))),
    # 经营风险(8)
    ("经营风险", "股权出质", lambda d: _risk_cell(d.get("EquityPledge"))),
    ("经营风险", "动产抵押", lambda d: _risk_cell(d.get("ChattelMortgage"))),
    ("经营风险", "欠税公告", lambda d: _risk_cell(d.get("TaxOweNotice"))),
    ("经营风险", "税务非正常户", lambda d: _risk_cell(d.get("TaxAbnormal"))),
    ("经营风险", "税收违法", lambda d: _risk_cell(d.get("TaxIllegal"))),
    ("经营风险", "清算信息", lambda d: _risk_cell(d.get("Liquidation"))),
    ("变更风险", "近期法定代表人/股东/注册资本变更", _f_change_recent),
]


def _data_of_2006(raw_2006: dict) -> dict:
    result = (raw_2006 or {}).get("Result") or {}
    if result.get("VerifyResult") != 1:
        return {}
    return result.get("Data") or {}


def _build_ctx(payloads: Dict[int, Any]) -> dict:
    """把 {apiCode: raw_resp} 组装成取值函数所需的 ctx。"""
    ctx = dict(_data_of_2006(payloads.get(2006)))
    ctx["_supp"] = {str(c): payloads.get(c) for c in (856, 739, 887, 888, 889) if c in payloads}
    return ctx


def load_payloads(company: str, snapshot_dir: str) -> Dict[int, dict]:
    """从 snapshots/ 读取该企业每个 apiCode 的最新快照(脱机,零成本)。

    与 snapshot.save 命名对齐: qcc_{apiCode}_{企业名}_{YYYYmmdd_HHMMSS}.json
    优先读 snapshot_dir/json/,若无则读 snapshot_dir 根(兼容旧落盘)。
    """
    safe = (company or "unknown").replace("/", "_")
    out: Dict[int, dict] = {}
    search_dirs = []
    json_dir = os.path.join(snapshot_dir, "json")
    if os.path.isdir(json_dir):
        search_dirs.append(json_dir)
    search_dirs.append(snapshot_dir)
    for code in (2006, 856, 739, 887, 888, 889):
        files = []
        for d in search_dirs:
            found = sorted(glob.glob(os.path.join(d, "qcc_%s_%s_*.json" % (code, safe))))
            if found:
                files = found
                break
        if not files:
            continue
        with open(files[-1], encoding="utf-8") as f:
            out[code] = json.load(f)
    return out


def build_report(company: str, payloads: Dict[int, Any], out_dir: str,
                 profile: Optional[dict] = None) -> str:
    """生成 qcc_{企业名称}_{时间}.xlsx,返回文件路径。

    profile: 可选,CompanyProfile 的关键字段,用于顶部「查询概要」分组;
             为 None 时(export 脱机场景)概要中风险/成本等填「—」。
    """
    ctx = _build_ctx(payloads or {})
    ts = time.strftime("%Y%m%d_%H%M%S")
    safe = (company or "unknown").replace("/", "_")
    filename = "qcc_%s_%s.xlsx" % (safe, ts)
    path = os.path.join(out_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "企业信息汇总"
    ws.append(["指标名", "指标值"])
    _style_header(ws, 2)

    # 顶部:查询概要分组
    _append_section(ws, "查询概要")
    summary_rows = [
        ("企业名称", company or "—"),
        ("查询时间", (profile or {}).get("query_date") or ts),
        ("风险等级", _level_zh((profile or {}).get("risk_level")) if profile else "—"),
        ("命中项数", str(len((profile or {}).get("hits", []))) if profile else "—"),
        ("调用接口", ", ".join(str(c) for c in (profile or {}).get("called_apis", [])) if profile else "—"),
        ("本次成本(元)", str((profile or {}).get("cost", "")) if profile else "—"),
    ]
    for name, val in summary_rows:
        ws.append([name, val])
    last = ws.max_row
    for c in (1, 2):
        ws.cell(row=last, column=c).border = BORDER

    # 指标分组(按 INDICATORS 顺序,相邻同分组合并成一个标题行)
    cur_group = None
    for group, name, fn in INDICATORS:
        if group != cur_group:
            _append_section(ws, group)
            cur_group = group
        try:
            val = fn(ctx) if ctx else "无数据(接口未开通)"
        except Exception:
            val = "无数据(接口未开通)"
        ws.append([name, val])
        r = ws.max_row
        for c in (1, 2):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.font = Font(name=FONT)

    _autofit(ws, [28, 90])
    ws.freeze_panes = "A2"

    os.makedirs(out_dir, exist_ok=True)
    wb.save(path)
    return path


def _level_zh(level: Any) -> str:
    return {"low": "低", "medium": "中", "high": "高", "unknown": "未知"}.get(level, str(level or "—"))


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 26


def _append_section(ws, title: str) -> None:
    ws.append([title, ""])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell = ws.cell(row=r, column=1)
    cell.value = title
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in (1, 2):
        ws.cell(row=r, column=c).border = BORDER


def _autofit(ws, widths: List[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
